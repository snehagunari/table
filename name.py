import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter as gcl
from openpyxl.utils.cell import coordinate_from_string as cfs

def extract_tables_from_sheet(sheet):
    """Extracts all tables from a given sheet and returns a dictionary of DataFrames."""
    tables = {}

    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
        header_row = [cell.value for cell in row]

        if any(header_row) and not all(v is None for v in header_row):  # Likely a valid table header
            first_col = cfs(row[0].coordinate)[0]
            first_row = row_idx

            # Detect last column safely
            last_col = gcl(sheet.max_column)  # Default to max column
            for col in range(1, sheet.max_column + 1):
                if sheet.cell(row=first_row + 1, column=col).value is None:
                    if col > 1:  # Ensure valid column index
                        last_col = gcl(col - 1)
                    break

            # Detect last row safely
            last_row = sheet.max_row
            for r in range(first_row + 1, sheet.max_row + 1):
                if all(sheet.cell(row=r, column=col).value is None for col in range(1, sheet.max_column + 1)):
                    last_row = r - 1
                    break

            # Extract table content
            rng = f"{first_col}{first_row}:{last_col}{last_row}"
            data_rows = [[cell.value for cell in row] for row in sheet[rng]]

            if data_rows:
                tables[f"Table_{first_row}"] = pd.DataFrame(data_rows[1:], columns=data_rows[0]).dropna(how="all")

    return tables

# File path to be processed
filename = r"C:\Users\sneha.gunari\Downloads\Untitled spreadsheet.xlsx"

try:
    wb = load_workbook(filename, data_only=True)
    all_tables = {}

    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        all_tables[sheet_name] = extract_tables_from_sheet(sheet)

    # Output extracted tables
    for sheet, tables in all_tables.items():
        for name, df in tables.items():
            if df.empty:
                print(f"Warning: '{name}' in '{sheet}' is empty!")
            else:
                print(f"\nExtracted {name} from '{sheet}'")
                print(df)
                print("----------------------------------\n")
except Exception as e:
    print(f"Error loading file: {e}")


